import openpyxl


def getrowcount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_row


def getcolumncount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_column


def readdata(file, sheetname, rowno ,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(row=rowno, column=columnno).value


def writedata(file, sheetname, rowno, columnno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row=rowno, column=columnno).value = data
    workbook.save()
